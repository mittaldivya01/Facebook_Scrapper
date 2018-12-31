#!/usr/bin/env python
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from openpyxl import Workbook
import openpyxl
import pymongo
from pymongo import MongoClient
import pprint
import sys ,time, getpass ,json
reload(sys)
sys.setdefaultencoding('utf8')



#Connectin to mongo db
client = MongoClient('localhost', 27017) 
db = client['psosm']	
collection = db['facebookdb']		


#login to facebook
driver = webdriver.Firefox()
def login():
	
	e= driver.find_element_by_id("m_login_email")
    	email = raw_input("Enter login email : ")
    	e.send_keys(email);
    	e = driver.find_element_by_name('pass')
    	try:
     		password = getpass.getpass('Enter password : ')
		e.send_keys(password); 
		e = driver.find_element_by_name('login')
    		e.click();
    	except Exception as error:
     		print('ERROR', error)
		driver.quit()



#global variables declaration
friend= []
fcount= 'null'

#Collect friends for user's facebookid
def findfriends(url):
	global fcount
	global friend
	fn , fl, fo= [], [], []
	try:
		driver.get(url);
		#get friend list
		if driver.find_element_by_xpath('/html/body/div/div/div[2]/div/div[1]/div[1]/div[4]/a[2]').text == "Friends":
			l=driver.find_element_by_xpath('/html/body/div/div/div[2]/div/div[1]/div[1]/div[4]/a[2]')
			link=l.get_attribute('href')
		else:
			#print 'no friends list available'
			friend.append("null")
			return
	except NoSuchElementException:
		return
		
	#finding count of friends
	try:
		driver.get(link);
		st = driver.find_element_by_xpath("/html/body/div/div/div[2]/div/div[1]/h3").text
 	except NoSuchElementException:
		friend.append("null")
		fcount= "null"
		return
	try:
		word= st.split()
		w= word[1].replace(',','')
		l= len(w)
		num= w[1:l-1]
		fcount= num
		print fcount
		num1 = int(num)
 	except ValueError:
		print "invalid friends count input"
		friend.append("null")
		fcount= st
		return
	
	#extract friends
	n=1
	try:
		for sn in range(1,num1 +1):
			sn1= str(n)
			try:
				fn.append(driver.find_element_by_css_selector('div.x:nth-child('+sn1+') > table:nth-child(1) > tbody:nth-child(1) > tr:nth-child(1) > td:nth-child(2) > a:nth-child(1)').text)

				try:
					fl.append(driver.find_element_by_css_selector('div.x:nth-child('+sn1+') > table:nth-child(1) > tbody:nth-child(1) > tr:nth-child(1) > td:nth-child(2) > a:nth-child(1)').get_attribute('href'))
				except NoSuchElementException:
					fl.append("null")
				try:
					fo.append(driver.find_element_by_css_selector('div.x:nth-child('+sn1+') > table:nth-child(1) > tbody:nth-child(1) > tr:nth-child(1) > td:nth-child(2) > div:nth-child(2) > span:nth-child(1)').text)
				except NoSuchElementException:
					fo.append("null")
				n +=1
			
			except NoSuchElementException:
				x= driver.find_element_by_id("m_more_friends")
				if driver.find_element_by_id("m_more_friends"):
					n= 1
					sn1= str(n)
				#check for more friends link: '
					link= x.find_element_by_tag_name('a').get_attribute('href')
					driver.get(link);
					fn.append(driver.find_element_by_css_selector('div.x:nth-child('+sn1+') > table:nth-child(1) > tbody:nth-child(1) > tr:nth-child(1) > td:nth-child(2) > a:nth-child(1)').text)	
							
					try:
						fl.append(driver.find_element_by_css_selector('div.x:nth-child('+sn1+') > table:nth-child(1) > tbody:nth-child(1) > tr:nth-child(1) > td:nth-child(2) > a:nth-child(1)').get_attribute('href'))
					except NoSuchElementException:
						fl.append("null")
					try:
						fo.append(driver.find_element_by_css_selector('div.x:nth-child('+sn1+') > table:nth-child(1) > tbody:nth-child(1) > tr:nth-child(1) > td:nth-child(2) > div:nth-child(2) > span:nth-child(1)').text)
					except NoSuchElementException:
						fo.append("null")
	
	except NoSuchElementException:
		print "total count not reached"		
	finally:
		print "check1"		
		friend = [{"fname": a, "flink": b, "others": c} for a, b, c in zip(fn, fl, fo)]	
		print friend
		
	print "check1"		
	friend = [{"fname": a, "flink": b, "others": c} for a, b, c in zip(fn, fl, fo)]	
	print friend
	return
		
#login function
try:
	driver.get("https://m.facebook.com/")
	login()
except NoSuchElementException:
	print "login error"

#workbook handeling
try:
	fi="/home/test/Inputfile.xlsx"
	wb2 = openpyxl.load_workbook(filename=fi)
	ws2 = wb2['Sheet1']
	last=ws2.max_row
except IOError:
	print "An error occured trying to read file"

#find friends for user profiles available 
for r in range(2,last):
	try:
		findfriends(ws2.cell(row=r,column=2).value)
		friendsnetwork={"name": ws2.cell(row=r,column=1).value, "facebookprofile": ws2.cell(row=r,column=2).value, "friends_count":fcount, "friends": friend}
		collection.insert_one(friendsnetwork)
	except NoSuchElementException:
		#print 'Profile not available'
		friendsnetwork={"name": ws2.cell(row=r,column=1).value, "facebookprofile": ws2.cell(row=r,column=2).value, "friends_count":fcount, "friends": 'null'}
		collection.insert_one(friendsnetwork)
		pass
	friend = []
	fcount= 'null'
	time.sleep(0.5)
driver.quit()
